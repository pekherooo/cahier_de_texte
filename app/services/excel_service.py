from io import BytesIO
from openpyxl import load_workbook

def generate_fiche_suivi(cours, enseignant, seances, modele_path='MODELE_FICHE_SUIVI.xlsx'):
    wb = load_workbook(modele_path)
    ws = wb.active

    ws['B5'] = enseignant.nom
    ws['G5'] = cours.nom
    ws['H5'] = f"{cours.volume_horaire} h prévues"

    start_row = 9
    for i, s in enumerate(seances):
        row = start_row + i
        ws[f'A{row}'] = i + 1
        ws[f'B{row}'] = s.date.strftime('%d/%m/%Y')
        ws[f'C{row}'] = s.heure_debut.strftime('%H:%M')
        ws[f'D{row}'] = s.heure_fin.strftime('%H:%M')
        ws[f'E{row}'] = round(s.duree.total_seconds() / 3600, 2)
        ws[f'F{row}'] = s.contenu
        ws[f'G{row}'] = "Oui" if s.validee else "Non"

    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)
    return excel_output
